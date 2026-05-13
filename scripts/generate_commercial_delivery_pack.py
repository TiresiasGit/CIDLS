"""
generate_commercial_delivery_pack.py
=====================================
商用請負納品文書パックを Excel (.xlsx) で生成するスクリプト。
CIDLS_PLATFORM_RESTORE.3 Step5 対応。
CIDLS_PLATFORM_RESTORE.2 「Excel出力: 全ドキュメント一覧・目次・属性・生成日時・
ファイルパス・進捗・品質・検収を .xlsx で出力」を実現する。

Usage:
    uv run python scripts/generate_commercial_delivery_pack.py

依存:
    openpyxl (uv が自動インストール)

# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""

import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment,
        Font,
        PatternFill,
        Border,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl が見つかりません。`uv add openpyxl` を実行してください。")
    sys.exit(1)

REPO_ROOT = Path(__file__).parent.parent
OUTPUT_DIR = REPO_ROOT / "reports" / "commercial_delivery"
TODAY = datetime.now().strftime("%Y-%m-%d")
NOW = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# =====================================================================
# 成果物マスターデータ
# =====================================================================
ARTIFACTS = [
    # (カテゴリ, ファイル名, 説明, ステータス, 優先度, 担当)
    ("CIDLSコア", "cidls_platform_overview.html", "CIDLSプラットフォーム全体像 (6ステップ・CAPDkA)", "DONE v2.2.0", "P1", "CIDLS"),
    ("CIDLSコア", "project_kanban.html",          "カンバンボード統合ハブ (SoT)",                    "DONE",       "P1", "CIDLS"),
    ("CIDLSコア", "graph_project_mindmap.html",   "DA表・スイムレーン・マインドマップ",              "DONE",       "P1", "CIDLS"),
    ("CIDLSコア", "STORY.html",                   "業務ストーリー v2.3.0 (6登場人物)",               "DONE",       "P1", "CIDLS"),
    ("SW納品",    "要求定義書.html",               "要求定義書 CIDLS-REQ-001 v1.0.0",                "DONE",       "P1", "CIDLS"),
    ("SW納品",    "コンセプトスライド.html",       "コンセプトスライド [SLIDE]準拠 7枚 SPA",          "DONE",       "P2", "CIDLS"),
    ("SW納品",    "要求仕様書.html",               "要求仕様書 (機能仕様・IF・データ仕様)",            "TODO",       "P2", "CIDLS"),
    ("SW納品",    "システム要件定義書.html",       "システム要件定義書",                              "TODO",       "P2", "CIDLS"),
    ("SW納品",    "基本設計書.html",               "基本設計書",                                      "TODO",       "P2", "CIDLS"),
    ("SW納品",    "詳細設計書.html",               "詳細設計書",                                      "TODO",       "P3", "CIDLS"),
    ("SW納品",    "DB設計書.html",                 "DB設計書 (DuckDB / Parquet スキーマ)",            "TODO",       "P3", "CIDLS"),
    ("SW納品",    "外部設計書.html",               "外部設計書 (IF設計書)",                           "TODO",       "P3", "CIDLS"),
    ("SW納品",    "結合テスト仕様書.html",         "結合テスト仕様書",                                "TODO",       "P3", "CIDLS"),
    ("SW納品",    "運用設計書.html",               "運用設計書",                                      "TODO",       "P3", "CIDLS"),
    ("SW納品",    "運用手順書.html",               "運用手順書",                                      "TODO",       "P3", "CIDLS"),
    ("SW納品",    "移行計画書.html",               "移行計画書",                                      "TODO",       "P3", "CIDLS"),
    ("SW納品",    "保守運用計画書.html",           "保守運用計画書",                                  "TODO",       "P3", "CIDLS"),
    ("SW納品",    "リリースノート.html",           "リリースノート",                                  "TODO",       "P3", "CIDLS"),
    ("スクリプト", "scripts/generate_cidls_platform_overview.py",   "overview自動更新スクリプト",  "DONE",  "P1", "CIDLS"),
    ("スクリプト", "scripts/generate_graph_project_mindmap.py",     "mindmap DA表更新スクリプト",  "DONE",  "P1", "CIDLS"),
    ("スクリプト", "scripts/generate_commercial_delivery_pack.py",  "Excelパック生成スクリプト",   "DONE",  "P1", "CIDLS"),
    ("スクリプト", "scripts/run_daily_self_evolution.ps1",          "日次11:00自律進化ループ",     "DONE",  "P1", "CIDLS"),
]

SW_DOCS = [
    ("要求定義書",       "共通フレーム2007 企画・要件定義フェーズ",   "DONE 2026-05-12 v1.0.0"),
    ("要求仕様書",       "共通フレーム2007 要件定義フェーズ詳細",     "TODO"),
    ("システム要件定義書", "共通フレーム2007 システム要件定義",       "TODO"),
    ("基本設計書",       "共通フレーム2007 開発フェーズ 設計",        "TODO"),
    ("詳細設計書",       "共通フレーム2007 開発フェーズ 詳細",        "TODO"),
    ("DB設計書",         "DuckDB / Parquet スキーマ定義",             "TODO"),
    ("外部設計書",       "IF設計書 (API / ファイル / 画面)",           "TODO"),
    ("結合テスト仕様書", "Given-When-Then 形式 [Q3]準拠",             "TODO"),
    ("運用設計書",       "日次ループ・監視・アラート設計",             "TODO"),
    ("運用手順書",       "setup / restore / devrag 手順",             "TODO"),
    ("移行計画書",       "データ移行・環境移行計画",                   "TODO"),
    ("保守運用計画書",   "月次レビュー・保守スコープ",                 "TODO"),
    ("リリースノート",   "バージョン・変更ログ・既知問題",             "TODO"),
]


def _thin_border():
    thin = Side(style="thin", color="B0BEC5")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, row: int, values: list[str], fill_hex: str = "1E3A5F") -> None:
    fill = _header_fill(fill_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _write_data_row(ws, row: int, values: list, status_col: int = -1) -> None:
    border = _thin_border()
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        # ステータス列の色付け
        if col == status_col:
            if "DONE" in str(val):
                cell.fill = _header_fill("1B5E20")
                cell.font = Font(color="A5D6A7", size=9)
            elif "TODO" in str(val):
                cell.fill = _header_fill("1A237E")
                cell.font = Font(color="90CAF9", size=9)
            elif "BLOCKED" in str(val):
                cell.fill = _header_fill("B71C1C")
                cell.font = Font(color="EF9A9A", size=9)


def build_cover_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "表紙"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:H2")
    title_cell = ws["B2"]
    title_cell.value = "CIDLS 商用請負 納品文書パック"
    title_cell.font = Font(bold=True, size=20, color="6366F1")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:H3")
    ws["B3"].value = f"Generated: {NOW}  |  Schema: v1.0.0  |  AGENTS.md準拠"
    ws["B3"].font = Font(size=10, color="94A3B8")
    ws["B3"].alignment = Alignment(horizontal="center")

    info = [
        ("プロジェクト名", "CIDLS (CI/DL System)"),
        ("生成日",         TODAY),
        ("バージョン",     "v1.0.0"),
        ("根拠",           "AGENTS.md [CIDLS_PLATFORM_RESTORE.2] / 共通フレーム2007"),
        ("成果物総数",     f"{len(ARTIFACTS)} 件"),
        ("SW納品ドキュメント", f"{len(SW_DOCS)} 種 (DONE={sum(1 for _,_,s in SW_DOCS if 'DONE' in s)})"),
    ]
    for i, (k, v) in enumerate(info, start=5):
        ws.cell(row=i, column=2, value=k).font = Font(bold=True, color="E2E8F0", size=10)
        ws.cell(row=i, column=3, value=v).font = Font(color="CBD5E1", size=10)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50
    ws.sheet_format.defaultRowHeight = 18
    ws.sheet_properties.tabColor = "6366F1"


def build_artifact_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("成果物一覧")
    ws.sheet_view.showGridLines = False
    headers = ["No", "カテゴリ", "ファイル名", "説明", "ステータス", "優先度", "担当", "ファイルサイズ"]
    _write_header_row(ws, 1, headers)
    _set_col_widths(ws, [5, 14, 42, 38, 16, 8, 10, 12])

    for i, (cat, fname, desc, status, prio, owner) in enumerate(ARTIFACTS, start=2):
        fpath = REPO_ROOT / fname
        size_str = f"{round(fpath.stat().st_size / 1024, 1)} KB" if fpath.exists() else "N/A"
        _write_data_row(ws, i, [i - 1, cat, fname, desc, status, prio, owner, size_str], status_col=5)

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "10B981"


def build_sw_docs_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("SW納品ドキュメント")
    ws.sheet_view.showGridLines = False
    headers = ["No", "ドキュメント名", "準拠フレームワーク", "ステータス", "ファイルパス", "備考"]
    _write_header_row(ws, 1, headers, fill_hex="0D47A1")
    _set_col_widths(ws, [5, 22, 36, 20, 36, 24])

    for i, (name, framework, status) in enumerate(SW_DOCS, start=2):
        fname = name + ".html"
        fpath = str(REPO_ROOT / fname)
        note = "作成済み" if "DONE" in status else "次回優先" if i <= 4 else "中期計画"
        _write_data_row(ws, i, [i - 1, name, framework, status, fpath, note], status_col=4)

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "2196F3"


def build_progress_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("進捗サマリ")
    ws.sheet_view.showGridLines = False

    done = sum(1 for *_, s, _, _ in ARTIFACTS if "DONE" in s)
    todo = len(ARTIFACTS) - done
    sw_done = sum(1 for _, _, s in SW_DOCS if "DONE" in s)
    sw_todo = len(SW_DOCS) - sw_done

    rows = [
        ("指標",                    "値"),
        ("成果物 DONE",             done),
        ("成果物 TODO/BLOCKED",     todo),
        ("成果物 合計",             len(ARTIFACTS)),
        ("SW納品ドキュメント DONE", sw_done),
        ("SW納品ドキュメント TODO", sw_todo),
        ("SW納品ドキュメント 合計", len(SW_DOCS)),
        ("生成日時",                NOW),
        ("次優先アクション",        "要求仕様書.html 作成 → generate_*.py 日次ループ統合"),
    ]
    _write_header_row(ws, 1, ["指標", "値"])
    _set_col_widths(ws, [30, 50])
    for i, (k, v) in enumerate(rows[1:], start=2):
        _write_data_row(ws, i, [k, v])

    ws.sheet_properties.tabColor = "F59E0B"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_DIR / f"商用請負納品文書パック_{TODAY}.xlsx"

    wb = openpyxl.Workbook()
    build_cover_sheet(wb)
    build_artifact_sheet(wb)
    build_sw_docs_sheet(wb)
    build_progress_sheet(wb)

    wb.save(output_file)
    size_kb = round(output_file.stat().st_size / 1024, 1)
    print(f"[SUCCESS] {output_file.name} を生成しました ({size_kb} KB)")
    print(f"[PATH]    {output_file}")


if __name__ == "__main__":
    main()
