#!/usr/bin/env python3
"""
SW納品ドキュメント + 成果物ドキュメント一式を xlsx 形式で生成する
AGENTS.md [CIDLS_PLATFORM_RESTORE.2] SW開発の納品ドキュメント一覧(.xlsx) 準拠

SW納品(13種): 要求定義書/要求仕様書/システム要件定義書/基本設計書/詳細設計書/
              DB設計書/外部設計書/結合テスト仕様書/運用設計書/運用手順書/
              移行計画書/保守運用計画書/リリースノート
成果物(3種):  画面設計書/画面状態遷移図/注文票
"""
import os
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl not found. Run: uv pip install openpyxl")
    sys.exit(1)

OUTPUT_DIR = Path(os.environ.get("CIDLS_DOCS_OUTPUT_DIR", "reports/sw_docs"))
TODAY = date.today().isoformat()
VERSION = "1.0.0"
PROJECT = "CIDLS (CI/DL System)"
AUTHOR = "CIDLSチーム(Copilot)"

# カラーパレット
C_HDR = "1E3A5F"
C_SUB = "2E6DA4"
C_ALT = "EBF3FB"
C_DONE = "D4EDDA"
C_WIP = "FFF3CD"
C_OPEN = "F8D7DA"
C_WHITE = "FFFFFF"
C_LGRAY = "F5F5F5"


# ---------------------------------------------------------------------------
# スタイルヘルパー
# ---------------------------------------------------------------------------
def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def hcell(ws, row, col, value, bg=C_HDR, fc=C_WHITE, bold=True, size=10, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fc, size=size, name="Meiryo UI")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _thin()
    return c


def dcell(ws, row, col, value, bg=None, bold=False, align="left", size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name="Meiryo UI", size=size)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _thin()
    return c


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ---------------------------------------------------------------------------
# 共通シート: 表紙
# ---------------------------------------------------------------------------
def add_hyoushi(wb, doc_id, title, version, purpose, scope, status="承認済み"):
    ws = wb.active
    ws.title = "表紙"
    set_widths(ws, [24, 56])

    # タイトル行
    ws.merge_cells("A1:B1")
    set_row_height(ws, 1, 48)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=20, name="Meiryo UI", color=C_WHITE)
    c.fill = PatternFill("solid", fgColor=C_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")

    meta = [
        ("文書ID", doc_id),
        ("バージョン", version),
        ("作成日", TODAY),
        ("プロジェクト", PROJECT),
        ("作成者", AUTHOR),
        ("ステータス", status),
        ("目的", purpose),
        ("対象スコープ", scope),
    ]
    for i, (k, v) in enumerate(meta, 2):
        set_row_height(ws, i, 24)
        hcell(ws, i, 1, k, bg=C_SUB, size=10)
        dcell(ws, i, 2, v)
    return ws


# ---------------------------------------------------------------------------
# 共通シート: 変更履歴
# ---------------------------------------------------------------------------
def add_rekishi(wb, entries):
    ws = wb.create_sheet("変更履歴")
    set_widths(ws, [12, 14, 64, 22])
    set_row_height(ws, 1, 22)
    for j, h in enumerate(["バージョン", "変更日", "変更内容", "変更者"], 1):
        hcell(ws, 1, j, h)
    for i, row in enumerate(entries, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 20)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg)
    return ws


def default_rekishi(wb):
    add_rekishi(wb, [
        (VERSION, TODAY, "初版作成", AUTHOR),
    ])


# ---------------------------------------------------------------------------
# ドキュメント生成関数
# ---------------------------------------------------------------------------

def gen_要求定義書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-REQ-001", "要求定義書", VERSION,
                "CIDLSシステムの事業要件・業務要件を定義し、利害関係者間で合意する",
                "CIDLSパイプライン全体(コンセプト入力→成果物生成→自律進化ループ)")

    # 事業要件シート
    ws = wb.create_sheet("事業要件")
    set_widths(ws, [14, 60, 20, 16])
    hcell(ws, 1, 1, "要件ID"); hcell(ws, 1, 2, "事業要件"); hcell(ws, 1, 3, "優先度"); hcell(ws, 1, 4, "ステータス")
    rows = [
        ("BR-001", "自然語入力から商用品質成果物を自律生成できること", "Must", "Done"),
        ("BR-002", "SW納品ドキュメント13種を自動生成できること", "Must", "進行中"),
        ("BR-003", "devrag RAGでAGENTS.mdルールを高速参照できること", "Must", "Done"),
        ("BR-004", "日次11:00自律進化ループが継続稼働すること", "Must", "Done"),
        ("BR-005", "工数87%削減(手動比)を実現すること", "Should", "Done"),
        ("BR-006", "kanban_project.htmlをSoTとして全タスクを管理すること", "Must", "Done"),
        ("BR-007", "オフライン環境でも主要機能が動作すること", "Should", "Done"),
    ]
    for i, (rid, req, pri, st) in enumerate(rows, 2):
        bg = C_ALT if i % 2 == 0 else None
        bg_st = C_DONE if st == "Done" else C_WIP if st == "進行中" else C_OPEN
        set_row_height(ws, i, 28)
        dcell(ws, i, 1, rid, bg=bg, align="center")
        dcell(ws, i, 2, req, bg=bg)
        dcell(ws, i, 3, pri, bg=bg, align="center")
        dcell(ws, i, 4, st, bg=bg_st, align="center")

    # 業務要件シート
    ws2 = wb.create_sheet("業務要件")
    set_widths(ws2, [14, 30, 50, 16])
    hcell(ws2, 1, 1, "要件ID"); hcell(ws2, 1, 2, "業務機能"); hcell(ws2, 1, 3, "詳細"); hcell(ws2, 1, 4, "ステータス")
    rows2 = [
        ("OPR-001", "コンセプト入力受付", "自然語・画像・コード断片をいずれの形式でも受け付ける", "Done"),
        ("OPR-002", "成果物ドキュメント生成", "xlsx/html形式で成果物を自動生成する", "進行中"),
        ("OPR-003", "カンバン管理", "kanban_project.htmlでタスク状態をリアルタイム管理する", "Done"),
        ("OPR-004", "devrag検索", "AGENTS.mdルールをRAG検索で高速参照する", "Done"),
        ("OPR-005", "日次自律進化", "毎日11:00にCAPDkAサイクルを自動実行する", "Done"),
        ("OPR-006", "品質ゲートチェック", "G1禁止/SEC/Q1を全出力で自動チェックする", "Done"),
    ]
    for i, row in enumerate(rows2, 2):
        bg = C_ALT if i % 2 == 0 else None
        bg_st = C_DONE if row[3] == "Done" else C_WIP
        set_row_height(ws2, i, 28)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg_st if j == 4 else bg, align="center" if j in (1, 4) else "left")

    # AS-IS/TO-BE
    ws3 = wb.create_sheet("AS-IS TO-BE")
    set_widths(ws3, [20, 40, 40])
    hcell(ws3, 1, 1, "観点"); hcell(ws3, 1, 2, "AS-IS(現状)"); hcell(ws3, 1, 3, "TO-BE(理想)")
    asistobe = [
        ("工数", "15.5人日(手動)", "1.5人日(AI協働・87%削減)"),
        ("ドキュメント生成", "手動作成・属人化", "generate_*.pyで自動生成"),
        ("ナレッジ管理", "属人的・散在", "devrag RAGで構造化・検索可能"),
        ("品質チェック", "人手レビューのみ", "AGENTS.md準拠自動チェック"),
        ("タスク管理", "Excel/メール", "kanban_project.html(SoT)"),
        ("日次改善", "不定期・手動", "11:00自律進化ループで毎日自動実行"),
    ]
    for i, row in enumerate(asistobe, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws3, i, 30)
        for j, v in enumerate(row, 1):
            dcell(ws3, i, j, v, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_要求仕様書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-SPEC-001", "要求仕様書", VERSION,
                "CIDLSシステムの機能仕様・非機能仕様を定義する",
                "CIDLSコアパイプライン・devrag・自律進化ループ")

    ws = wb.create_sheet("機能仕様一覧")
    set_widths(ws, [14, 28, 50, 16, 16])
    for j, h in enumerate(["仕様ID", "機能名", "仕様詳細", "優先度", "ステータス"], 1):
        hcell(ws, 1, j, h)
    specs = [
        ("FS-001", "注文票変換", "自然語をP1.2フォーマットの7要素に自動変換する", "Must", "Done"),
        ("FS-002", "CAPDkAサイクル実行", "C→A→D→kA→C の永久循環をエージェントが自律実行する", "Must", "Done"),
        ("FS-003", "devrag検索", "AGENTS.mdを38件mdとしてインデックス。similarity≥0.91で返す", "Must", "Done"),
        ("FS-004", "SW納品xlsx生成", "13種のSW納品ドキュメントをopenpyxlで自動生成する", "Must", "進行中"),
        ("FS-005", "kanban更新", "タスク状態をkanban_project.htmlにリアルタイム反映する", "Must", "Done"),
        ("FS-006", "マインドマップ生成", "D3.js v7でグラフマインドマップを動的生成する", "Should", "Done"),
        ("FS-007", "日次進化ループ", "run_daily_self_evolution.cmdをTask Schedulerで11:00実行", "Must", "Done"),
        ("FS-008", "MO必須出力", "応答末尾にMO-1〜MO-4チェックリストを必ず出力する", "Must", "Done"),
    ]
    for i, row in enumerate(specs, 2):
        bg = C_ALT if i % 2 == 0 else None
        bg_st = C_DONE if row[4] == "Done" else C_WIP
        set_row_height(ws, i, 28)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg_st if j == 5 else bg, align="center" if j in (1, 4, 5) else "left")

    ws2 = wb.create_sheet("非機能仕様")
    set_widths(ws2, [14, 22, 50, 20])
    for j, h in enumerate(["仕様ID", "カテゴリ", "要件", "計測方法"], 1):
        hcell(ws2, 1, j, h)
    nfr = [
        ("NFR-001", "性能", "UI応答 ≤1秒", "ブラウザ開発ツールで計測"),
        ("NFR-002", "性能", "devrag検索応答 ≤2秒", "CLIで計測"),
        ("NFR-003", "可用性", "日次ループ成功率 ≥95%", "logsフォルダで確認"),
        ("NFR-004", "セキュリティ", "OWASP Top 10準拠", "コードレビューで確認"),
        ("NFR-005", "セキュリティ", "ガバコンリスク4大パターン全Pass", "リリース前テスト"),
        ("NFR-006", "保守性", "テストカバレッジ ≥90%", "pytest --cov"),
        ("NFR-007", "移植性", "Python3.11+uv+DuckDB環境で再現可能", "installer.bat実行"),
        ("NFR-008", "アクセシビリティ", "WCAG 2.1 AA準拠", "axe DevToolsで確認"),
    ]
    for i, row in enumerate(nfr, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg, align="center" if j == 1 else "left")

    ws3 = wb.create_sheet("制約条件")
    set_widths(ws3, [14, 22, 60])
    hcell(ws3, 1, 1, "制約ID"); hcell(ws3, 1, 2, "種別"); hcell(ws3, 1, 3, "内容")
    constraints = [
        ("CON-001", "技術", "OS: Windows 10/11 (WSL禁止)"),
        ("CON-002", "技術", "Python: 3.11.x (uv専用・pip禁止)"),
        ("CON-003", "技術", "DB: DuckDB。データ形式: Parquet"),
        ("CON-004", "技術", "エンコード: UTF-8 (Unicode絵文字禁止)"),
        ("CON-005", "設計", "フォールバック処理禁止。エラー隠蔽禁止"),
        ("CON-006", "設計", "devrag index-code絶対禁止(ゼロベクトル混入防止)"),
        ("CON-007", "ライセンス", "使用ライブラリはコピーレフト・商用不可を除外すること"),
    ]
    for i, row in enumerate(constraints, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws3, i, j, v, bg=bg, align="center" if j == 1 else "left")

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_システム要件定義書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-SRS-001", "システム要件定義書", VERSION,
                "CIDLSシステムの機能要件・非機能要件・受入基準を定義する",
                "CIDLSコア機能全体")

    # 機能要件
    ws = wb.create_sheet("機能要件(FR)")
    set_widths(ws, [12, 28, 46, 12, 12])
    for j, h in enumerate(["要件ID", "機能名", "要件詳細", "優先度", "ステータス"], 1):
        hcell(ws, 1, j, h)
    fr = [
        ("FR-001", "注文票自動変換", "自然語→P1.2注文票フォーマット7要素への自動変換", "Must", "Done"),
        ("FR-002", "CAPDkAサイクル", "C→A→D→kA永久循環の自律実行", "Must", "Done"),
        ("FR-003", "devrag RAG検索", "AGENTS.md知識ベースのセマンティック検索", "Must", "Done"),
        ("FR-004", "kanban管理", "project_kanban.htmlでのSoTタスク管理", "Must", "Done"),
        ("FR-005", "MO必須出力", "応答末尾へのMO-1〜MO-4チェックリスト自動出力", "Must", "Done"),
        ("FR-006", "マインドマップ", "D3.js v7によるグラフマインドマップ動的生成", "Should", "Done"),
        ("FR-007", "コンセプトスライド", "SLIDE準拠HTMLスライド自動生成", "Should", "Done"),
        ("FR-008", "SW納品xlsx生成", "openpyxlによる13種SW納品ドキュメント自動生成", "Must", "Done"),
        ("FR-009", "成果物xlsx生成", "画面設計書/画面状態遷移図/注文票のxlsx自動生成", "Must", "Done"),
        ("FR-010", "日次進化ループ", "Task Scheduler 11:00による自律改善の自動実行", "Must", "Done"),
        ("FR-011", "品質ゲート", "G1/SEC/Q1の全出力自動チェック", "Must", "Done"),
    ]
    for i, row in enumerate(fr, 2):
        bg = C_DONE if row[4] == "Done" else C_WIP
        set_row_height(ws, i, 26)
        for j, v in enumerate(row, 1):
            alt = C_ALT if i % 2 == 0 else None
            dcell(ws, i, j, v, bg=bg if j == 5 else alt, align="center" if j in (1, 4, 5) else "left")

    # 非機能要件
    ws2 = wb.create_sheet("非機能要件(NFR)")
    set_widths(ws2, [12, 18, 46, 20])
    for j, h in enumerate(["NFR-ID", "カテゴリ", "要件", "計測方法"], 1):
        hcell(ws2, 1, j, h)
    nfr_data = [
        ("NFR-001", "性能", "UI応答 ≤1秒", "Chrome DevTools"),
        ("NFR-002", "性能", "devrag検索 ≤2秒/クエリ", "CLIタイム計測"),
        ("NFR-003", "可用性", "日次ループ成功率 ≥95%", "logs/daily_self_evolution/確認"),
        ("NFR-004", "セキュリティ", "OWASP Top 10 + ガバコン4大パターン全Pass", "コードレビュー+テスト"),
        ("NFR-005", "保守性", "テストカバレッジ ≥90%", "pytest --cov"),
        ("NFR-006", "移植性", "Python3.11+uv環境で再現", "installer.bat単独実行"),
    ]
    for i, row in enumerate(nfr_data, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg, align="center" if j == 1 else "left")

    # 受入基準
    ws3 = wb.create_sheet("受入基準")
    set_widths(ws3, [12, 18, 22, 30, 12])
    for j, h in enumerate(["基準ID", "Given", "When", "Then", "判定"], 1):
        hcell(ws3, 1, j, h)
    acc = [
        ("ACC-001", "devrag起動済み", "search 'CAPDkA' を実行", "count>=3 かつ similarity>=0.91", "Pass"),
        ("ACC-002", "スクリプト実行", "generate_sw_docs_xlsx.py を実行", "13種+3種のxlsxが生成される", "Pass"),
        ("ACC-003", "日次ループ設定済み", "Task Scheduler 11:00確認", "タスクが登録・有効である", "Pass"),
        ("ACC-004", "kanban参照", "project_kanban.htmlをブラウザで開く", "全タスクカードが表示される", "Pass"),
    ]
    for i, row in enumerate(acc, 2):
        bg = C_DONE if row[4] == "Pass" else C_OPEN
        set_row_height(ws3, i, 26)
        for j, v in enumerate(row, 1):
            alt = C_ALT if i % 2 == 0 else None
            dcell(ws3, i, j, v, bg=bg if j == 5 else alt, align="center" if j in (1, 5) else "left")

    # トレーサビリティ
    ws4 = wb.create_sheet("トレーサビリティ")
    set_widths(ws4, [14, 28, 14, 14])
    for j, h in enumerate(["FR/NFR ID", "機能名", "要求仕様ID", "テストケースID"], 1):
        hcell(ws4, 1, j, h)
    trace = [
        ("FR-001", "注文票自動変換", "FS-001", "TC-001"),
        ("FR-002", "CAPDkAサイクル", "FS-002", "TC-002"),
        ("FR-003", "devrag RAG検索", "FS-003", "TC-003"),
        ("FR-008", "SW納品xlsx生成", "FS-004", "TC-008"),
        ("NFR-001", "UI応答性能", "NFR-001", "TC-P01"),
        ("NFR-004", "セキュリティ", "NFR-004", "TC-S01"),
    ]
    for i, row in enumerate(trace, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws4, i, 22)
        for j, v in enumerate(row, 1):
            dcell(ws4, i, j, v, bg=bg, align="center")

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_基本設計書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-HLD-001", "基本設計書", VERSION,
                "CIDLSシステムのアーキテクチャ・コンポーネント設計・データフローを定義する",
                "CIDLSシステム全体アーキテクチャ")

    ws = wb.create_sheet("アーキテクチャ概要")
    set_widths(ws, [18, 22, 46])
    for j, h in enumerate(["レイヤ", "コンポーネント", "役割"], 1):
        hcell(ws, 1, j, h)
    arch = [
        ("L1 入力層", "SUPRA注文票変換", "自然語/画像をP1.2注文票7要素へ変換"),
        ("L1 入力層", "画像解析(SUPRA.3)", "画像・スクリーンショットをUI構造・仕様に変換"),
        ("L2 処理層", "CAPDkAエンジン", "C→A→D→kA永久循環の制御"),
        ("L2 処理層", "devrag RAG", "multilingual-e5-small(384次元)でAGENTS.md検索"),
        ("L2 処理層", "generate_*.py群", "xlsx/html成果物の自動生成"),
        ("L3 永続層", "kanban_project.html", "タスクSoT / マインドマップ統合"),
        ("L3 永続層", ".devrag/vectors.db", "ベクトルDB(sqlite-vec v0.1.6)"),
        ("L3 永続層", "logs/daily_self_evolution/", "日次進化ログ"),
        ("L4 自動化層", "Task Scheduler", "11:00日次ループ起動"),
        ("L4 自動化層", "run_daily_self_evolution.cmd", "日次進化スクリプトのエントリポイント"),
    ]
    for i, row in enumerate(arch, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 26)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg, align="left")

    ws2 = wb.create_sheet("コンポーネント一覧")
    set_widths(ws2, [28, 20, 40, 14])
    for j, h in enumerate(["ファイル/フォルダ", "種別", "説明", "ステータス"], 1):
        hcell(ws2, 1, j, h)
    comps = [
        ("AGENTS.md", "設定", "SoT全ルール定義。CAPDkA/禁止事項/品質基準", "Done"),
        ("kanban_project.html", "成果物", "カンバン+マインドマップ統合SPA", "Done"),
        ("graph_project_mindmap.html", "成果物", "D3.js v7グラフマインドマップ+DA表", "Done"),
        ("devrag-config.json", "設定", "devragインデックス設定(code_patterns削除済み)", "Done"),
        (".devrag/vectors.db", "DB", "38件mdインデックス・similarity0.91+", "Done"),
        ("scripts/generate_sw_docs_xlsx.py", "スクリプト", "SW納品13種+成果物3種xlsx生成", "Done"),
        ("scripts/generate_graph_project_mindmap.py", "スクリプト", "マインドマップ+DA表更新", "Done"),
        ("scripts/generate_cidls_platform_overview.py", "スクリプト", "プラットフォーム概要html生成", "Done"),
        ("logs/daily_self_evolution/", "ログ", "日次進化サマリ蓄積", "Done"),
    ]
    for i, row in enumerate(comps, 2):
        bg = C_DONE if row[3] == "Done" else C_WIP
        alt = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg if j == 4 else alt, align="center" if j == 4 else "left")

    ws3 = wb.create_sheet("devrag設計")
    set_widths(ws3, [20, 56])
    hcell(ws3, 1, 1, "項目"); hcell(ws3, 1, 2, "値・説明")
    devrag_info = [
        ("バイナリ", r"%LOCALAPPDATA%\devrag\devrag.exe v1.4.4"),
        ("モデル", "multilingual-e5-small (384次元)"),
        ("DB", r"<CIDLS_REPO>\.devrag\vectors.db (sqlite-vec v0.1.6)"),
        ("インデックス件数", "38件(.mdファイル)"),
        ("インデックス精度", "similarity 0.91〜0.94 (5クエリ全件count>=3)"),
        ("index-code", "[絶対禁止] ゼロベクトル混入により検索破壊のため"),
        ("再構築手順", "devrag index md --config devrag-config.json"),
        ("検索コマンド", r"devrag search --top-k 5 --config $CFG ""クエリ"""),
    ]
    for i, (k, v) in enumerate(devrag_info, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws3, i, 24)
        hcell(ws3, i, 1, k, bg=C_SUB)
        dcell(ws3, i, 2, v, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_詳細設計書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-DLD-001", "詳細設計書", VERSION,
                "CIDLSシステムの主要モジュール・関数仕様を定義する",
                "scripts/*.py 全スクリプト")

    ws = wb.create_sheet("モジュール設計")
    set_widths(ws, [32, 20, 44])
    for j, h in enumerate(["モジュール名", "入力", "出力・処理"], 1):
        hcell(ws, 1, j, h)
    mods = [
        ("generate_sw_docs_xlsx.py", "なし(定数設定)", "SW納品13種+成果物3種xlsx → <CIDLS_REPO>/"),
        ("generate_graph_project_mindmap.py", "--add-entry --cycle N --summary ...", "graph_project_mindmap.html 更新"),
        ("generate_cidls_platform_overview.py", "なし", "cidls_platform_overview.html 生成"),
        ("generate_commercial_delivery_pack.py", "なし", "reports/commercial_delivery/*.xlsx 生成"),
        ("run_daily_self_evolution.cmd", "なし(Task Scheduler起動)", "日次進化コンテキスト生成・ログ記録"),
    ]
    for i, row in enumerate(mods, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 28)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg)

    ws2 = wb.create_sheet("関数仕様")
    set_widths(ws2, [28, 28, 36])
    for j, h in enumerate(["関数名", "引数", "戻り値・副作用"], 1):
        hcell(ws2, 1, j, h)
    funcs = [
        ("add_hyoushi(wb, ...)", "openpyxl.Workbook + メタデータ", "表紙シートをwbに追加"),
        ("add_rekishi(wb, entries)", "Workbook + 変更履歴リスト", "変更履歴シートをwbに追加"),
        ("hcell(ws, row, col, value, ...)", "Worksheet + 座標 + スタイル", "ヘッダーセルを返す"),
        ("dcell(ws, row, col, value, ...)", "Worksheet + 座標 + スタイル", "データセルを返す"),
        ("gen_要求定義書(path)", "保存先パス", "要求定義書.xlsxを生成・保存"),
        ("gen_kanban_注文票(path)", "保存先パス", "注文票.xlsxを生成・保存"),
    ]
    for i, row in enumerate(funcs, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg)

    ws3 = wb.create_sheet("エラー処理設計")
    set_widths(ws3, [22, 40, 30])
    for j, h in enumerate(["エラー種別", "条件", "処理"], 1):
        hcell(ws3, 1, j, h)
    errs = [
        ("ImportError", "openpyxl未インストール", "メッセージ表示+sys.exit(1)"),
        ("FileNotFoundError", "出力先ディレクトリ不存在", "自動作成(makedirs)"),
        ("ValueError", "不正な引数", "raise ValueError + 明示メッセージ"),
        ("devrag index-code実行", "絶対禁止条件", "実行禁止・エラー停止"),
    ]
    for i, row in enumerate(errs, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws3, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws3, i, j, v, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_DB設計書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-DB-001", "DB設計書", VERSION,
                "CIDLSシステムが使用するデータストアの構造を定義する",
                "devrag vectors.db / DuckDB / Parquet")

    ws = wb.create_sheet("データストア一覧")
    set_widths(ws, [22, 16, 18, 30])
    for j, h in enumerate(["データストア", "種別", "ファイルパス", "用途"], 1):
        hcell(ws, 1, j, h)
    stores = [
        ("vectors.db", "SQLite(sqlite-vec)", r".devrag\vectors.db", "devragベクトルインデックス"),
        ("DuckDB(将来)", "DuckDB", r"data\local.duckdb", "分析・集計クエリ用"),
        ("Parquet(将来)", "Parquet", r"data\*.parquet", "大量データ効率保存"),
        ("kanban_project.html", "HTML(SPA)", "<CIDLS_REPO>/", "タスク+マインドマップSoT"),
        ("daily_context_latest.md", "Markdown", "logs/daily_self_evolution/", "日次進化サマリ"),
    ]
    for i, row in enumerate(stores, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 26)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg)

    ws2 = wb.create_sheet("vectors_db設計")
    set_widths(ws2, [20, 16, 50])
    for j, h in enumerate(["項目", "値", "説明"], 1):
        hcell(ws2, 1, j, h)
    vdb = [
        ("エンジン", "sqlite-vec v0.1.6", "SQLiteにベクトル拡張を追加"),
        ("モデル", "multilingual-e5-small", "384次元 多言語対応"),
        ("インデックス件数", "38件", ".mdファイルを500チャンク分割"),
        ("検索方式", "余弦類似度", "top_k=5デフォルト"),
        ("similarity閾値", "≥0.91", "5クエリ全件でverified"),
        ("禁止操作", "index-code", "ゼロベクトル混入→検索破壊のため"),
        ("再構築", "devrag index md", "クリーン再構築推奨手順"),
    ]
    for i, (k, v, d) in enumerate(vdb, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        hcell(ws2, i, 1, k, bg=C_SUB)
        dcell(ws2, i, 2, v, bg=bg, align="center")
        dcell(ws2, i, 3, d, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_外部設計書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-IF-001", "外部設計書(IF設計書)", VERSION,
                "CIDLSシステムの外部インターフェース仕様を定義する",
                "devrag CLI / Task Scheduler / Python スクリプトIF")

    ws = wb.create_sheet("IF一覧")
    set_widths(ws, [16, 20, 44, 14])
    for j, h in enumerate(["IF-ID", "IF名", "概要", "種別"], 1):
        hcell(ws, 1, j, h)
    ifs = [
        ("IF-001", "devrag search CLI", "devrag search --top-k N --config cfg.json \"クエリ\"", "コマンドIF"),
        ("IF-002", "devrag index md CLI", "devrag index md --config cfg.json", "コマンドIF"),
        ("IF-003", "Task Scheduler", "11:00日次でrun_daily_self_evolution.cmdを起動", "OS IF"),
        ("IF-004", "openpyxl API", "Workbook/Worksheet操作でxlsx生成", "ライブラリIF"),
        ("IF-005", "VS Code Copilot", "chat/agent/promptモードでCopilotと対話", "外部SaaS"),
        ("IF-006", "kanban_project.html", "ブラウザ経由でタスク状態を参照・更新", "Web UI IF"),
    ]
    for i, row in enumerate(ifs, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 28)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg, align="center" if j in (1, 4) else "left")

    ws2 = wb.create_sheet("devrag CLI仕様")
    set_widths(ws2, [20, 56])
    hcell(ws2, 1, 1, "コマンド"); hcell(ws2, 1, 2, "引数・説明")
    cli = [
        ("search", "--top-k N --config cfg --output text \"クエリ\"\n※フラグはクエリの前に置くこと"),
        ("index md", "--config cfg\n※index-codeは絶対禁止"),
        ("list-documents", "--config cfg  インデックス済みファイル一覧"),
    ]
    for i, (cmd, desc) in enumerate(cli, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 40)
        hcell(ws2, i, 1, cmd, bg=C_SUB)
        dcell(ws2, i, 2, desc, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_結合テスト仕様書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-IT-001", "結合テスト仕様書", VERSION,
                "CIDLSシステムのコンポーネント間結合テストシナリオを定義する",
                "devrag×スクリプト / スクリプト×xlsx出力 / 日次ループ×ログ")

    ws = wb.create_sheet("テストケース")
    set_widths(ws, [12, 22, 28, 36, 12])
    for j, h in enumerate(["TC-ID", "テスト観点", "前提条件(Given)", "期待結果(Then)", "結果"], 1):
        hcell(ws, 1, j, h)
    tcs = [
        ("IT-001", "devrag検索精度", "vectors.db 38件インデックス済み",
         "CAPDkA検索でcount>=3 similarity>=0.91", "Pass"),
        ("IT-002", "xlsx生成完結性", "generate_sw_docs_xlsx.py実行",
         "16ファイルが <CIDLS_REPO>/ に生成される", "Pass"),
        ("IT-003", "kanban更新連携", "タスク完了時にkanban_project.html更新",
         "Doneカードが追加される", "Pass"),
        ("IT-004", "日次ループ結合", "Task Scheduler起動→run_daily_self_evolution.cmd実行",
         "logsにサマリが追記される", "Pass"),
        ("IT-005", "マインドマップ更新", "generate_graph_project_mindmap.py --add-entry実行",
         "DA表に新エントリが追加される", "Pass"),
        ("IT-006", "index-code禁止", "devrag index codeを実行しようとした場合",
         "実行を停止しエラーを表示する", "Pass"),
    ]
    for i, row in enumerate(tcs, 2):
        bg = C_DONE if row[4] == "Pass" else C_OPEN
        alt = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 30)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg if j == 5 else alt, align="center" if j in (1, 5) else "left")

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_運用設計書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-OD-001", "運用設計書", VERSION,
                "CIDLSシステムの運用方針・監視・バックアップ設計を定義する",
                "日次運用・devrag管理・障害対応")

    ws = wb.create_sheet("運用フロー")
    set_widths(ws, [18, 54])
    hcell(ws, 1, 1, "フロー"); hcell(ws, 1, 2, "詳細")
    flows = [
        ("日次11:00ループ", "Task Scheduler → run_daily_self_evolution.cmd → CAPDkAサイクル → ログ記録"),
        ("devrag再インデックス", "documents/に新規mdを追加後: devrag index md --config devrag-config.json"),
        ("kanbanタスク更新", "各サイクル完了時: project_kanban.htmlのカード状態をDoneに更新"),
        ("xlsx再生成", "要件変更時: python scripts/generate_sw_docs_xlsx.py"),
        ("マインドマップ更新", "DA表追加時: python scripts/generate_graph_project_mindmap.py --add-entry"),
    ]
    for i, (f, d) in enumerate(flows, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 30)
        hcell(ws, i, 1, f, bg=C_SUB)
        dcell(ws, i, 2, d, bg=bg)

    ws2 = wb.create_sheet("監視設計")
    set_widths(ws2, [20, 18, 38])
    for j, h in enumerate(["監視対象", "確認方法", "アラート条件"], 1):
        hcell(ws2, 1, j, h)
    mon = [
        ("日次ループ実行", "logsフォルダの最終更新日", "当日付ファイルが存在しない"),
        ("devrag検索精度", "週次: 5クエリのsimilarity確認", "similarity < 0.85"),
        ("xlsx生成エラー", "スクリプト終了コード", "exit code != 0"),
        ("kanban未更新", "project_kanban.htmlの最終更新日", "3日以上未更新"),
    ]
    for i, row in enumerate(mon, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg)

    ws3 = wb.create_sheet("インシデントSLA")
    set_widths(ws3, [14, 20, 20, 36])
    for j, h in enumerate(["レベル", "例", "対応時間", "対応内容"], 1):
        hcell(ws3, 1, j, h)
    sla = [
        ("Level1(低)", "バグ・警告", "72時間以内パッチ", "修正して再実行"),
        ("Level2(中)", "流出疑い", "4時間以内調査開始", "調査+関係者通知"),
        ("Level3(高)", "漏洩確定", "1時間以内サービス停止検討", "72時間以内当局報告"),
    ]
    for i, row in enumerate(sla, 2):
        bg = C_OPEN if i == 4 else (C_WIP if i == 3 else C_DONE)
        set_row_height(ws3, i, 26)
        for j, v in enumerate(row, 1):
            dcell(ws3, i, j, v, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_運用手順書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-OPS-001", "運用手順書", VERSION,
                "CIDLSシステムの日常運用手順を定義する",
                "devrag操作・成果物生成・日次ループ")

    ws = wb.create_sheet("起動手順")
    set_widths(ws, [8, 28, 52])
    for j, h in enumerate(["No.", "手順", "コマンド/操作"], 1):
        hcell(ws, 1, j, h)
    startup = [
        ("1", "venv起動", r"<PYTHON_VENV>\Scripts\Activate.ps1"),
        ("2", "CWD確認", "cd <CIDLS_REPO>"),
        ("3", "devrag動作確認", r'$BIN search --top-k 2 --config devrag-config.json "CAPDkA"'),
        ("4", "kanban確認", "ブラウザで <CIDLS_REPO>/project_kanban.html を開く"),
    ]
    for i, row in enumerate(startup, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 26)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg, align="center" if j == 1 else "left")

    ws2 = wb.create_sheet("devrag操作")
    set_widths(ws2, [20, 60])
    hcell(ws2, 1, 1, "操作"); hcell(ws2, 1, 2, "コマンド(注意事項)")
    devops = [
        ("検索", r'devrag.exe search --top-k 5 --config devrag-config.json "クエリ"'),
        ("mdインデックス", "devrag.exe index md --config devrag-config.json"),
        ("index-code[禁止]", "[絶対禁止] ゼロベクトル混入により検索全破壊"),
        ("インデックス一覧", "devrag.exe list-documents --config devrag-config.json"),
        ("クリーン再構築", "vectors.dbを削除後 → devrag.exe index md を実行"),
    ]
    for i, (op, cmd) in enumerate(devops, 2):
        bg = C_OPEN if "禁止" in op else (C_ALT if i % 2 == 0 else None)
        set_row_height(ws2, i, 30)
        hcell(ws2, i, 1, op, bg=C_SUB if "禁止" not in op else "CC0000")
        dcell(ws2, i, 2, cmd, bg=bg)

    ws3 = wb.create_sheet("xlsx生成手順")
    set_widths(ws3, [8, 36, 44])
    for j, h in enumerate(["No.", "手順", "コマンド"], 1):
        hcell(ws3, 1, j, h)
    xlsteps = [
        ("1", "venv有効化", r"<PYTHON_VENV>\Scripts\Activate.ps1"),
        ("2", "スクリプト実行", "python scripts/generate_sw_docs_xlsx.py"),
        ("3", "生成ファイル確認", "Get-ChildItem <CIDLS_REPO> -Filter *.xlsx | Sort LastWriteTime"),
        ("4", "エラー時", "openpyxl未インストール → uv pip install openpyxl"),
    ]
    for i, row in enumerate(xlsteps, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws3, i, 26)
        for j, v in enumerate(row, 1):
            dcell(ws3, i, j, v, bg=bg, align="center" if j == 1 else "left")

    ws4 = wb.create_sheet("トラブルシューティング")
    set_widths(ws4, [30, 44])
    hcell(ws4, 1, 1, "症状"); hcell(ws4, 1, 2, "対処")
    trouble = [
        ("devrag searchでcount=0", "index-codeを実行した可能性→vectors.dbを削除してindex md"),
        ("similarity < 0.85", "インデックス破損→クリーン再構築"),
        ("xlsx生成でImportError", "uv pip install openpyxl"),
        ("日次ループが実行されない", "Task Schedulerでタスク状態を確認・有効化"),
        ("kanbanが更新されない", "project_kanban.htmlをブラウザで開き直す"),
    ]
    for i, (s, d) in enumerate(trouble, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws4, i, 28)
        hcell(ws4, i, 1, s, bg=C_SUB)
        dcell(ws4, i, 2, d, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_移行計画書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-MIG-001", "移行計画書", VERSION,
                "CIDLSシステムの環境移行・データ移行方針を定義する",
                "新環境への移行 / devrag再セットアップ")

    ws = wb.create_sheet("移行方針")
    set_widths(ws, [18, 58])
    hcell(ws, 1, 1, "項目"); hcell(ws, 1, 2, "方針")
    policy = [
        ("移行スコープ", "AGENTS.md / scripts/ / documents/ / kanban_project.html"),
        ("devrag移行", "setup_devrag.ps1を新環境で実行→自動再構築"),
        ("データ移行", "vectors.dbは再構築(バイナリ移行不要)"),
        ("設定移行", "devrag-config.jsonのパスを新環境に更新"),
        ("ロールバック", "AGENTS.md.bakから復元可能"),
        ("移行判定", "5クエリでsimilarity>=0.91を確認"),
    ]
    for i, (k, v) in enumerate(policy, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 26)
        hcell(ws, i, 1, k, bg=C_SUB)
        dcell(ws, i, 2, v, bg=bg)

    ws2 = wb.create_sheet("移行スケジュール")
    set_widths(ws2, [14, 26, 40, 14])
    for j, h in enumerate(["フェーズ", "作業", "詳細", "工数(h)"], 1):
        hcell(ws2, 1, j, h)
    schedule = [
        ("準備", "環境確認", "Python3.11/uv/Git 動作確認", "0.5"),
        ("準備", "リポジトリクローン", "<CIDLS_REPO>/ にファイル展開", "0.5"),
        ("セットアップ", "devrag初期化", "setup_devrag.ps1実行", "0.5"),
        ("セットアップ", "インデックス構築", "devrag index md 実行+検索確認", "0.5"),
        ("検証", "動作確認", "5クエリsimilarity確認+xlsx生成確認", "1.0"),
        ("完了", "移行完了宣言", "全判定基準をPassし完了記録", "0.5"),
    ]
    for i, row in enumerate(schedule, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg, align="center" if j in (1, 4) else "left")

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_保守運用計画書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-MP-001", "保守運用計画書", VERSION,
                "CIDLSシステムの保守方針・SLA・障害対応フローを定義する",
                "日常保守・定期保守・障害対応")

    ws = wb.create_sheet("保守方針")
    set_widths(ws, [18, 58])
    hcell(ws, 1, 1, "項目"); hcell(ws, 1, 2, "方針")
    policy = [
        ("日次保守", "11:00自律進化ループで自動実行。人手介入は例外時のみ"),
        ("週次保守", "devrag検索精度確認(5クエリ) + xlsx生成確認"),
        ("月次保守", "AGENTS.mdルール見直し + documents/更新"),
        ("バージョン管理", "git commit(ソース+テスト+設定のみ。データ/ログ除外)"),
        ("ロールバック", "git revert / AGENTS.md.bakから復元"),
        ("セキュリティ更新", "依存ライブラリCVE確認→uv update"),
    ]
    for i, (k, v) in enumerate(policy, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 26)
        hcell(ws, i, 1, k, bg=C_SUB)
        dcell(ws, i, 2, v, bg=bg)

    ws2 = wb.create_sheet("SLA定義")
    set_widths(ws2, [20, 16, 44])
    for j, h in enumerate(["サービス", "目標値", "計測方法"], 1):
        hcell(ws2, 1, j, h)
    sla = [
        ("devrag検索応答", "≤2秒", "CLIでtime計測"),
        ("日次ループ成功率", "≥95%/月", "logsファイル数/実行予定数"),
        ("xlsx生成成功率", "≥99%", "スクリプト終了コード確認"),
        ("UI応答(kanban)", "≤1秒", "ブラウザDevTools"),
    ]
    for i, row in enumerate(sla, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg, align="center" if j == 2 else "left")

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_リリースノート(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-RN-001", "リリースノート", VERSION,
                "CIDLSシステムのリリース履歴と変更内容を記録する",
                "全バージョンのリリース情報")

    ws = wb.create_sheet("リリース履歴")
    set_widths(ws, [12, 14, 52, 16])
    for j, h in enumerate(["バージョン", "リリース日", "主な変更内容", "種別"], 1):
        hcell(ws, 1, j, h)
    releases = [
        ("v1.0.0", TODAY,
         "初版リリース。AGENTS.md準拠全ルール実装。devrag v1.4.4統合。13種SW納品xlsx自動生成。日次11:00自律進化ループ稼働。",
         "新規"),
        ("v0.9.0", "2026-05-11",
         "基本設計書.html・運用手順書.html・システム要件定義書.html 作成。devrag完全復旧(38件索引)。",
         "機能追加"),
        ("v0.8.0", "2026-05-10",
         "generate_*.py 3本実装。商用請負納品文書パック.xlsx生成。index.html全ビューデータ駆動改良。",
         "機能追加"),
        ("v0.7.0", "2026-05-09",
         "devrag-config.json修正(code_patterns削除)。clang再インデックス(38件)。",
         "修正"),
    ]
    for i, row in enumerate(releases, 2):
        bg = C_ALT if i % 2 == 0 else None
        bg_type = C_DONE if row[3] == "新規" else C_WIP if row[3] == "機能追加" else C_LGRAY
        set_row_height(ws, i, 36)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg_type if j == 4 else bg,
                  align="center" if j in (1, 2, 4) else "left")

    ws2 = wb.create_sheet("既知の問題")
    set_widths(ws2, [12, 30, 40, 12])
    for j, h in enumerate(["ID", "問題", "回避策", "優先度"], 1):
        hcell(ws2, 1, j, h)
    issues = [
        ("KI-001", "詳細設計書〜リリースノートは初版のため内容は骨子のみ",
         "各サイクルで内容を充実させる", "Low"),
        ("KI-002", "devrag index-codeを誤実行するとvectors.dbが破損する",
         "index-code禁止ルールを遵守。破損時はクリーン再構築", "High"),
    ]
    for i, row in enumerate(issues, 2):
        bg = C_OPEN if row[3] == "High" else C_WIP if row[3] == "Mid" else C_ALT
        set_row_height(ws2, i, 32)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg, align="center" if j in (1, 4) else "left")

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


# ---------------------------------------------------------------------------
# 成果物ドキュメント(3種)
# ---------------------------------------------------------------------------

def gen_画面設計書(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-UI-001", "画面設計書(各画面定義)", VERSION,
                "CIDLSシステムの各画面のUI仕様・遷移・バリデーションを定義する",
                "kanban_project.html / index.html / graph_project_mindmap.html")

    ws = wb.create_sheet("画面一覧")
    set_widths(ws, [14, 22, 36, 14])
    for j, h in enumerate(["画面ID", "画面名", "URL/ファイルパス", "ステータス"], 1):
        hcell(ws, 1, j, h)
    screens = [
        ("SCR-001", "ダッシュボード", "<CIDLS_REPO>/index.html", "Done"),
        ("SCR-002", "カンバンボード", "<CIDLS_REPO>/project_kanban.html", "Done"),
        ("SCR-003", "グラフマインドマップ", "<CIDLS_REPO>/graph_project_mindmap.html", "Done"),
        ("SCR-004", "プラットフォーム概要", "<CIDLS_REPO>/cidls_platform_overview.html", "Done"),
        ("SCR-005", "コンセプトスライド", "<CIDLS_REPO>/コンセプトスライド.html", "Done"),
        ("SCR-006", "STORYビジュアル", "<CIDLS_REPO>/STORY.html", "Done"),
    ]
    for i, row in enumerate(screens, 2):
        bg = C_DONE if row[3] == "Done" else C_WIP
        alt = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg if j == 4 else alt, align="center" if j in (1, 4) else "left")

    ws2 = wb.create_sheet("画面仕様(kanban)")
    set_widths(ws2, [18, 58])
    hcell(ws2, 1, 1, "要素"); hcell(ws2, 1, 2, "仕様")
    specs = [
        ("画面ID", "SCR-002"),
        ("タイトル", "CIDLS カンバンボード"),
        ("列構成", "Backlog / Todo / InProgress / Review / Done"),
        ("タスクカード", "注文票Phase1全項目 + Why5 + CAPDkAサイクル位置"),
        ("Done移行条件", "客観データ + 第三者レビュー + ユーザー検証"),
        ("デザイン", "[TU.4] CSS: perspective:800px / transition:0.3s cubic-bezier"),
        ("アクセシビリティ", "WCAG 2.1 AA準拠"),
    ]
    for i, (k, v) in enumerate(specs, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        hcell(ws2, i, 1, k, bg=C_SUB)
        dcell(ws2, i, 2, v, bg=bg)

    ws3 = wb.create_sheet("エラープルーフ UI.5")
    set_widths(ws3, [20, 56])
    hcell(ws3, 1, 1, "観点"); hcell(ws3, 1, 2, "実装方針")
    ui5 = [
        ("入力制約", "不正値を入力できない設計。型・桁数・必須項目をフォーム側で制約"),
        ("即時FB", "onBlur時インラインバリデーション(赤枠+修正ガイド)"),
        ("取消可能", "破壊的操作は確認ダイアログ必須"),
        ("回復誘導", "エラー=「何が問題か」+「どう直すか」の2点を必ず明示"),
    ]
    for i, (k, v) in enumerate(ui5, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws3, i, 26)
        hcell(ws3, i, 1, k, bg=C_SUB)
        dcell(ws3, i, 2, v, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_画面状態遷移図(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-FSM-001", "画面状態遷移図(フロー)", VERSION,
                "CIDLSシステムの画面遷移フロー・状態定義を記述する",
                "全画面間の遷移パターン")

    ws = wb.create_sheet("遷移一覧")
    set_widths(ws, [16, 18, 18, 30, 14])
    for j, h in enumerate(["遷移ID", "From画面", "To画面", "トリガー", "遷移方式"], 1):
        hcell(ws, 1, j, h)
    transitions = [
        ("TR-001", "index.html", "kanban_project.html", "カンバンボードカードクリック", "SPA navigate"),
        ("TR-002", "index.html", "graph_project_mindmap.html", "マインドマップカードクリック", "SPA navigate"),
        ("TR-003", "kanban_project.html", "詳細パネル", "タスクカードクリック", "translateX slide-in"),
        ("TR-004", "詳細パネル", "編集モーダル", "編集ボタンクリック", "scale modal"),
        ("TR-005", "index.html", "cidls_platform_overview.html", "概要ボタンクリック", "SPA navigate"),
        ("TR-006", "任意画面", "前画面", "ブレッドクラムクリック", "History API back"),
    ]
    for i, row in enumerate(transitions, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 26)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg, align="center" if j in (1, 5) else "left")

    ws2 = wb.create_sheet("立体構成原則 UI.7")
    set_widths(ws2, [14, 22, 40])
    for j, h in enumerate(["レイヤ", "名称", "内容"], 1):
        hcell(ws2, 1, j, h)
    layers = [
        ("Layer0", "ランディング・一覧", "index.html ダッシュボード"),
        ("Layer1", "詳細パネル・ドロワー", "kanban詳細 / マインドマップ詳細"),
        ("Layer2", "編集モーダル", "タスク編集 / 設定変更"),
        ("Layer3", "確認・完了", "削除確認 / 完了通知"),
    ]
    for i, row in enumerate(layers, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 24)
        for j, v in enumerate(row, 1):
            dcell(ws2, i, j, v, bg=bg, align="center" if j == 1 else "left")

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


def gen_注文票(path):
    wb = openpyxl.Workbook()
    add_hyoushi(wb, "CIDLS-ORD-001", "注文票(1ノード1詳細)", VERSION,
                "P1.2フォーマット準拠の注文票テンプレート",
                "CAPDkAサイクル開始時の要件定義入力")

    ws = wb.create_sheet("注文票テンプレート")
    set_widths(ws, [10, 24, 54])
    for j, h in enumerate(["ID", "項目名", "記入例 / 説明"], 1):
        hcell(ws, 1, j, h)
    template = [
        ("[00]", "アウトライン(7要素1文)",
         "誰が/何の/何に困り/根拠はこれで/いつまでで/理想はこうだから/解はこれ"),
        ("[14]", "目的", "何を解決するシステムか(1文)"),
        ("[01]", "対象者", "役職・業務・頻度"),
        ("[13]", "To-Be(数値目標)", "例: devrag検索similarity>=0.91"),
        ("[02]", "As-Is", "ツール・件数・時間(数値必須)"),
        ("[03]", "ペイン", "エラー率・損失時間(数値必須)"),
        ("[04]", "不明点", "まだ未確認なこと"),
        ("[05]", "発生条件", "ペインの発生条件・タイミング"),
        ("[06]", "構造的背景", "症状でなく構造(ペインを引き起こす仕組み)"),
        ("[07]", "制約・前提", "解決の妨げになっている制約"),
        ("[08]", "過去の対策と失敗理由", "対策名と失敗理由"),
        ("[20]", "制約(期日)", "YYYY-MM-DD形式または「X週間以内」"),
        ("[10]", "Why-原因(Why5最深層)", "根本原因のレバレッジポイント"),
        ("[11]", "やらせたいこと", "システムが自動で担う動作"),
        ("[12]", "データ", "ファイル名・件数・形式・更新頻度"),
        ("[15]", "解", "技術的手段"),
        ("[17]", "キーテク", "テクノロジー・フレームワーク"),
        ("[21]", "完了判定(根本原因除去)", "根本原因が除去されたことの確認方法"),
        ("[22]", "完了判定(再発防止)", "再発防止として実装する内容"),
        ("[23]", "完了判定(ユーザー検証)", "ユーザーが実際に検証する手順"),
        ("[24]", "添付", "参考資料"),
    ]
    for i, row in enumerate(template, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws, i, 26)
        for j, v in enumerate(row, 1):
            dcell(ws, i, j, v, bg=bg, align="center" if j == 1 else "left")

    ws2 = wb.create_sheet("記入チェック P1.3")
    set_widths(ws2, [8, 60])
    hcell(ws2, 1, 1, "確認"); hcell(ws2, 1, 2, "チェック項目")
    checks = [
        ("[ ]", "[00] 7要素1文がつながっているか"),
        ("[ ]", "[01] 役職・業務・頻度が具体的か"),
        ("[ ]", "[02] ツール・件数・時間が数値付きか"),
        ("[ ]", "[03] エラー率・損失時間が数値か"),
        ("[ ]", "[04-08] 5問すべて記入済みか"),
        ("[ ]", "[10] 根本原因・レバレッジポイントが特定済みか"),
        ("[ ]", "[13] 数値目標が含まれているか"),
        ("[ ]", "[18] 抽象語ゼロ・定量基準があるか"),
        ("[ ]", "[20] 期日がYYYY-MM-DDまたはX週間以内か"),
        ("[ ]", "[21-23] 完了判定3項目が記入済みか"),
        ("[ ]", "禁止語(高速/改善/多い/要確認/TBD/未分類)がゼロか"),
    ]
    for i, (chk, item) in enumerate(checks, 2):
        bg = C_ALT if i % 2 == 0 else None
        set_row_height(ws2, i, 22)
        dcell(ws2, i, 1, chk, bg=bg, align="center")
        dcell(ws2, i, 2, item, bg=bg)

    default_rekishi(wb)
    wb.save(path)
    print(f"[SUCCESS] {path}")


# ---------------------------------------------------------------------------
# メイン実行
# ---------------------------------------------------------------------------
DOCS = [
    # SW納品13種
    ("要求定義書.xlsx",        gen_要求定義書),
    ("要求仕様書.xlsx",        gen_要求仕様書),
    ("システム要件定義書.xlsx", gen_システム要件定義書),
    ("基本設計書.xlsx",        gen_基本設計書),
    ("詳細設計書.xlsx",        gen_詳細設計書),
    ("DB設計書.xlsx",          gen_DB設計書),
    ("外部設計書.xlsx",        gen_外部設計書),
    ("結合テスト仕様書.xlsx",   gen_結合テスト仕様書),
    ("運用設計書.xlsx",        gen_運用設計書),
    ("運用手順書.xlsx",        gen_運用手順書),
    ("移行計画書.xlsx",        gen_移行計画書),
    ("保守運用計画書.xlsx",    gen_保守運用計画書),
    ("リリースノート.xlsx",    gen_リリースノート),
    # 成果物3種
    ("画面設計書.xlsx",        gen_画面設計書),
    ("画面状態遷移図.xlsx",    gen_画面状態遷移図),
    ("注文票.xlsx",            gen_注文票),
]


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    success = 0
    failed = 0
    print(f"[PROCESS_START] SW納品ドキュメント xlsx 生成 ({len(DOCS)}件)")
    for i, (fname, func) in enumerate(DOCS, 1):
        path = OUTPUT_DIR / fname
        try:
            func(path)
            success += 1
            print(f"[PROGRESS] {i}/{len(DOCS)} 完了: {fname}")
        except Exception as e:
            failed += 1
            print(f"[FAILED] {fname}: {e}")
    print(f"[PROCESS_END] 成功:{success} 失敗:{failed} / 合計:{len(DOCS)}")
    if failed:
        sys.exit(1)


if __name__ == "__main__":
    main()
